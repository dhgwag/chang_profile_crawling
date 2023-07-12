from selenium import webdriver
from selenium.common.exceptions import WebDriverException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from config import *

# Chromedriver 없을 시 처음에는 자동으로 설치합니다.
chromedriver_path = r'C:\workspace\chromedriver.exe'


class Chang:
    def __init__(self):
        # Login
        self.login_id = id

        self.driver = None
        self.action = None

    def run_driver(self):
        try:
            self.driver = webdriver.Chrome()
            self.action = ActionChains(self.driver)
            return True
        except WebDriverException:
            self.driver = webdriver.Chrome(ChromeDriverManager().install())
            self.tabs = self.driver.window_handles
            self.action = ActionChains(self.driver)
            return False

    def chang_login(self):
            try:
                self.driver.set_window_size(500, 1024)
                self.driver.get('https://chang.spartacodingclub.kr/profiles')
                WebDriverWait(self.driver, 60).until(
                    EC.presence_of_element_located((By.CLASS_NAME, 'css-1nzrmep'))
                )
                self.driver.find_element(By.CLASS_NAME, "css-1nzrmep").click()
                WebDriverWait(self.driver, 60).until(
                    EC.presence_of_element_located((By.NAME, 'loginId'))
                )
                self.driver.find_element(By.NAME, 'loginId').send_keys(str(kakao_id))
                self.driver.find_element(By.NAME, 'password').send_keys(str(kakao_pw))

                self.driver.find_element(By.XPATH, "/html/body/div/div/div/main/article/div/div/form/div[4]/button[1]").click()

                WebDriverWait(self.driver, 600).until(
                    EC.presence_of_element_located((By.CLASS_NAME, 'css-m6pqt6'))
                )
            except Exception as e:
                print(e)
                print("에러발생")
                pass

    def crawl_by_name(self, name):
        return ILLEGAL_CHARACTERS_RE.sub(r'', self.driver.find_element(By.NAME, name).get_attribute('value'))

    def crawl(self):
        # 엑셀파일 쓰기
        write_wb = Workbook()

        # 이름이 있는 시트를 생성
        write_ws = write_wb.create_sheet('창 참여자 명단')
        write_ws = write_wb.active
        write_ws.append(['이름', '연락처', '이메일', '직무', '거주지', 'MBTI', '나의 성향', '학력', '이력', '스킬', '단기간 목표',
                         '대표 희망 여부', '아이디어 유/무', '외부 동업자 유/무', '원하는 팀 멤버의 성향', '한 줄 어필', '창업 관심 분야',
                         '희망 시기', '창업이 하고 싶은 이유', '실제 경험 공유'])
        idx = 0
        while True:
            section_list = self.driver.find_elements(By.XPATH, "/html/body/div[1]/section[1]/section[3]/section")
            print('-----------------------------------------' + str(len(section_list)))
            print(idx)
            if idx < len(section_list):
                section_list[idx].click()

                WebDriverWait(self.driver, 60).until(
                    EC.presence_of_element_located((By.CLASS_NAME, 'css-1h9nnjl'))
                )

                name = self.driver.find_element(By.CLASS_NAME, "css-1h9nnjl").get_attribute('innerText')
                info = self.driver.find_elements(By.CLASS_NAME, "css-1iqjuad")
                phone = info[0].get_attribute('innerText').replace('\n', '').replace('연락처', '')
                email = info[1].get_attribute('innerText').replace('\n', '').replace('이메일', '')
                name = ILLEGAL_CHARACTERS_RE.sub(r'', name)
                phone = ILLEGAL_CHARACTERS_RE.sub(r'', phone)
                email = ILLEGAL_CHARACTERS_RE.sub(r'', email)

                job = self.crawl_by_name("job")
                address = self.crawl_by_name("address")
                mbti = self.crawl_by_name("mbti")
                character = self.crawl_by_name("character")

                education = self.crawl_by_name("education")
                career = self.crawl_by_name("career")
                skill = self.crawl_by_name("skill")

                team_building = self.driver.find_elements(By.CLASS_NAME, "css-171g49t")
                if len(team_building) == 4:
                    short_target = team_building[0].get_attribute('innerText').replace('\n', '').replace(' ', '')
                    want_ceo = team_building[1].get_attribute('innerText')
                    has_idea = team_building[2].get_attribute('innerText')
                    external_cofounder = team_building[3].get_attribute('innerText')
                else:
                    short_target = '사이드프로젝트'
                    want_ceo = team_building[0].get_attribute('innerText')
                    has_idea = team_building[1].get_attribute('innerText')
                    external_cofounder = team_building[2].get_attribute('innerText')
                short_target = ILLEGAL_CHARACTERS_RE.sub(r'', short_target)
                want_ceo = ILLEGAL_CHARACTERS_RE.sub(r'', want_ceo)
                has_idea = ILLEGAL_CHARACTERS_RE.sub(r'', has_idea)
                external_cofounder = ILLEGAL_CHARACTERS_RE.sub(r'', external_cofounder)

                idealMemberDescription = self.crawl_by_name("idealMemberDescription")
                appealMention = self.crawl_by_name("appealMention")

                interest = self.crawl_by_name("interest")
                desiredTime = self.crawl_by_name("desiredTime")
                reasonWhy = self.crawl_by_name("reasonWhy")
                experience = self.crawl_by_name("experience")

                print(f"{name}    {phone}    {email}")
                # print(f"{job}    {address}    {mbti}    {character}")
                # print(f"{education}    {career}    {skill}")
                # print(f"{short_target}    {want_ceo}    {has_idea}    {external_cofounder}")
                # print(f"{idealMemberDescription}    {appealMention}")
                # print(f"{interest}    {desiredTime}    {reasonWhy}    {experience}")

                write_ws.append([name, phone, email, job, address, mbti, character, education, career, skill, short_target,
                                 want_ceo, has_idea, external_cofounder, idealMemberDescription, appealMention, interest,
                                 desiredTime, reasonWhy, experience])
                if short_target == '전업창업':
                    fill_color = '86E57F'
                elif short_target == '사이드창업':
                    fill_color = 'FFFFCE'
                elif short_target == '사이드프로젝트':
                    fill_color = 'FFC7CE'
                for i in range(1, write_ws.max_column + 1):
                    write_ws.cell(row=idx + 2, column=i).fill = PatternFill(start_color=fill_color,
                                                                            fill_type='solid')  ## 배경색 추가
                    # 행 단위로 추가

                self.driver.back()

                WebDriverWait(self.driver, 60).until(
                    EC.presence_of_element_located((By.CLASS_NAME, 'css-18z6k3f'))
                )
                idx += 1
            else:
                break

        write_wb.save("창_참가자.xlsx")

    def run(self):
        self.run_driver()
        self.chang_login()
        self.crawl()

if __name__ == "__main__":
    chang_crawler = Chang()
    chang_crawler.run()